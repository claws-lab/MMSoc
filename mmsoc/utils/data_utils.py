import ast
import json
import os
import os.path as osp
import platform
import re
import shutil
import sys
import traceback
from os import path as osp
from typing import List

import pandas as pd
import torch
from datasets import Features, Value, DatasetDict
from openpyxl import load_workbook

sys.path.append(os.path.abspath('.'))


def pad_tensor(tensor: torch.Tensor, max_size: int):
    tensor = torch.nn.functional.pad(tensor, (0, max_size - tensor.shape[-1]), 'constant', 0)
    return tensor


def pad_tensors(tensors: List[torch.Tensor]):
    max_size = max([t.shape[-1] for t in tensors])

    padded_input_ids = [pad_tensor(t, max_size) for t in tensors]

    padded_input_ids = torch.cat(padded_input_ids, dim=0)

    return padded_input_ids


def save(answer_li: list, answers_file: str, existing_results=None):
    with pd.ExcelWriter(answers_file) as writer:
        results_df = pd.DataFrame(answer_li)

        if existing_results is not None:
            results_df = pd.concat([existing_results, results_df], ignore_index=True)

        results_df.to_excel(writer, index=False)


def truncate_input(text: str, max_num_words: int = 256):
    try:
        text = re.sub(r'\s+', ' ', re.sub(r"\n+", " ", text))
        words = text.split(" ")
        words = words[:max_num_words]
        text = " ".join(words)
        if not any(text.endswith(punctuation) for punctuation in ".,:;?!|"):
            text += '.'

    except:
        traceback.print_exc()
        text = ""

    return text


def keep_ascii_chars(s):
    return re.sub(r'[^\x00-\x7F]+', '', s)


def get_device_for_platform():
    if platform.system() == "Windows":
        device = "cuda:0"
    elif platform.system() == "Darwin":
        device = "mps:0"
    elif platform.system() == "Linux":
        device = "cuda"
    else:
        raise ValueError("Unknown platform.")

    return device


def load_existing_results(answers_file: str):
    existing_results = None

    # Load questions if exists
    if osp.exists(answers_file):
        existing_results = pd.read_excel(answers_file, engine='openpyxl')
        START = len(existing_results)
        print(f"Starting from {START}th question")

    return existing_results


def get_gpt4v_generated_data_file_path(dataset_name: str, split: str):
    data_file = (f'outputs/GPT4V/image_desc_GPT4V_{dataset_name}_'
                 f'{split}.json')

    return data_file


def get_modified_time_of_file(path):
    import datetime, pathlib
    model_metadata = pathlib.Path(path).stat()
    mtime = datetime.datetime.fromtimestamp(model_metadata.st_mtime)
    ctime = datetime.datetime.fromtimestamp(model_metadata.st_ctime)
    print(f"\t{osp.basename(path)}: modified {mtime} | created {ctime}")
    return mtime, ctime


def load_jsonl(path):
    # Initialize an empty list to store the parsed JSON objects
    data = []

    # Open the JSONL file and parse each line as a JSON object
    with open(path, 'r') as jsonl_file:
        for idx, line in enumerate(jsonl_file):
            try:
                json_obj = json.loads(line)
                data.append(json_obj)
            except json.JSONDecodeError:
                print(f"Error parsing line {idx}")
                # print(line)

    # Create a DataFrame from the list of JSON objects
    df = pd.DataFrame(data)

    return df


def load_data(dataset_name: str, split: str, data_dir: str):
    print(f"Loading data from {dataset_name} (split: {split})")

    # Sentiment Analysis: Memotion
    if dataset_name in ["memotion"]:
        all_data = pd.read_excel(osp.join(data_dir, "Social", "memotion_dataset_7k", "labels.xlsx"), index_col=0)

        train_size = int(len(all_data) * 0.8)
        val_size = int(len(all_data) * 0.1)

        new_df = []

        for split in ["train", "val", "test"]:

            if split == "train":
                df = all_data[:train_size]

            elif split == "val":
                df = all_data[train_size:(train_size + val_size)]

            elif split == "test":
                df = all_data[(train_size + val_size):]

            os.makedirs(osp.join(data_dir, "Social", "MMSoc_memotion", split), exist_ok=True)

            for idx, row in df.iterrows():

                source_file = osp.join(data_dir, "Social", "memotion_dataset_7k", "images", row['image'])
                target_file = osp.join(data_dir, "Social", "MMSoc_memotion", split, row['image'])
                if os.path.exists(source_file):
                    shutil.copy2(source_file, target_file)

                    new_row = row.copy()
                    new_row['image'] = osp.join(split, row['image'])
                    new_row['split'] = split
                    new_df.append(new_row)

                # row['image'] = osp.join(row['image'])

            # if row['label'] == "neutral":
            #     df.loc[idx, 'label'] = 0
            # elif row['label'] == "positive":
            #     df.loc[idx, 'label'] = 1
            # elif row['label'] == "negative":
            #     df.loc[idx, 'label'] = 2

        # Create a new DataFrame from the collected rows
        all_data = pd.DataFrame(new_df).reset_index(drop=True)
        all_data.rename({"image": "file_name"}, axis=1, inplace=True)

        # Save the new DataFrame to JSON lines format
        all_data.to_json(osp.join(data_dir, "Social", "MMSoc_memotion", "metadata.jsonl"), orient="records", lines=True)

        from datasets import load_dataset

        features = Features({
            'file_name': Value('string'),
            'text_ocr': Value('string'),
            'text_corrected': Value('string'),
            'humor': Value('string'),
            'sarcasm': Value('string'),
            'offensive': Value('string'),
            'motivational': Value('string'),
            'sentiment': Value('string'),
            'split': Value('string'),
        })

        metadata_file = os.path.join(data_dir, "Social", "MMSoc_memotion", "metadata.jsonl")

        dataset = load_dataset(
            'json',
            data_files=metadata_file,
            features=features,
        )

        # Split the dataset into train, validation, and test sets
        train_dataset = dataset.filter(lambda x: x['split'] == 'train')['train']
        val_dataset = dataset.filter(lambda x: x['split'] == 'val')['val']
        test_dataset = dataset.filter(lambda x: x['split'] == 'test')['test']

        dataset_dict = DatasetDict({
            'train': train_dataset,
            'val': val_dataset,
            'test': test_dataset
        })

        image_directory_name = osp.join(data_dir, "images")


    # Misinformation: FakeNewsNet
    elif dataset_name in ["politi", "gossip"]:
        df = pd.read_csv(osp.join(data_dir, "Social", "FakeNewsNet_Multimodal", f"{dataset_name}_{split}.csv"))
        df = df.rename({"Unnamed: 0": "id"}, axis=1).set_index("id")

        df['label'] = 1 - df['label'].astype(int)

        label_sum = df['label'].sum()

        image_directory_name = osp.join(data_dir, "Images", f"{dataset_name}_{split}")

        print("Sum of labels:", label_sum)
        print("Number of examples:", len(df))


    # Hatespeech

    elif dataset_name in ["hatefulmemes"]:
        df = load_jsonl(osp.join(data_dir, "Social", "HatefulMemes", f"{split}.jsonl"))
        # print("TODO: Loading only the first 200 entries")
        # df = df[:200]
        df = df.rename({"img": "image"}, axis=1)
        image_directory_name = osp.join(data_dir, "img")

    elif dataset_name.startswith("YouTube"):
        df = pd.read_excel(osp.join(data_dir, "Social", dataset_name, "YouTube_video_metadata.xlsx"))
        image_directory_name = osp.join(data_dir, "Social", dataset_name, "img")
        df['topics'] = df['topics'].apply(ast.literal_eval)
        assert isinstance(df['topics'].iloc[0], list)

        train_size = int(len(df) * 0.8)

        if split == "train":
            df = df[:train_size]

        elif split == "test":
            df = df[train_size:]

        else:
            print("Unknown split")


    elif dataset_name == "toxic_comments":
        df = pd.read_csv(osp.join(data_dir, "Social", "ToxicCommentsDataset", "train.csv"))
        image_directory_name = None

    else:
        raise NotImplementedError

    df = df.reset_index(drop=True)

    return df


def check_sheet_exists(file_path, sheet_name):
    workbook = load_workbook(file_path)
    return sheet_name in workbook.sheetnames


def get_image_path(args, line: pd.Series) -> str:
    if args.dataset_name in ["politi", "gossip"]:
        image_file = osp.join(args.data_dir, "Social", "FakeNewsNet_Multimodal", "Images", f"{args.dataset_name}"
                                                                                           f"_{args.split}",
                              line["image"])

    elif args.dataset_name in ["memotion"]:
        image_file = osp.join(args.data_dir, "Social", "memotion_dataset_7k", "images", line["image"])

    elif args.dataset_name in ["hatefulmemes"]:
        image_file = osp.join(args.data_dir, "Social", "HatefulMemes", line["image"])

    elif args.dataset_name.startswith("YouTube"):
        image_file = osp.join(args.data_dir, "Social", args.dataset_name, "img", f"{line['id']}.jpg")

    else:
        raise NotImplementedError(f"Unknown dataset: {args.dataset_name}")

    return image_file
